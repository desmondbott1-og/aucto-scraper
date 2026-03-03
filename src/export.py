"""Export scraped data from SQLite to Excel."""

import json
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from . import config
from .db import ensure_db, get_full_export_data

logger = logging.getLogger(__name__)

HEADERS = [
    "Title",
    "URL",
    "Price",
    "Currency",
    "Seller Name",
    "Location",
    "Primary Category",
    "Subcategory",
    "Image URLs",
    "Item Details (Core Specs)",
    "All Detail Images",
]


def export_to_excel(output_path: str | None = None) -> str:
    """Export all data to an Excel file. Returns the output path."""
    ensure_db()
    path = output_path or str(config.EXPORT_PATH)
    data = get_full_export_data()

    if not data:
        logger.warning("No data to export!")
        return path

    wb = Workbook()
    ws = wb.active
    ws.title = "Aucto Listings"

    # Style header
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Write data rows
    for row_idx, item in enumerate(data, 2):
        image_urls = item.get("image_urls") or "[]"
        if isinstance(image_urls, str):
            try:
                image_urls = json.loads(image_urls)
            except json.JSONDecodeError:
                image_urls = []

        core_specs = item.get("core_specifications") or "{}"
        if isinstance(core_specs, str):
            try:
                core_specs = json.loads(core_specs)
            except json.JSONDecodeError:
                core_specs = {}

        all_images = item.get("all_image_urls") or "[]"
        if isinstance(all_images, str):
            try:
                all_images = json.loads(all_images)
            except json.JSONDecodeError:
                all_images = []

        # Format core specs as readable text
        specs_text = "\n".join(f"{k}: {v}" for k, v in core_specs.items()) if core_specs else ""

        values = [
            item.get("title", ""),
            item.get("item_url", ""),
            item.get("price", ""),
            item.get("currency", "USD"),
            item.get("seller_name", ""),
            item.get("location", ""),
            item.get("primary_category", ""),
            item.get("subcategory", ""),
            "\n".join(image_urls) if image_urls else "",
            specs_text,
            "\n".join(all_images) if all_images else "",
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Auto-adjust column widths (approximate)
    col_widths = [40, 60, 12, 8, 30, 40, 25, 30, 50, 50, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    config.DATA_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    logger.info("Exported %d rows to %s", len(data), path)
    return path


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    export_to_excel()
